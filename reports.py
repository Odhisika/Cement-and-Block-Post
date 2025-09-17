import tkinter as tk
from tkinter import messagebox, filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from datetime import datetime
import csv
import logging
from gui_utils import create_button_frame, create_card_frame

# Excel export functionality
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")

# Safe ToolTip import
try:
    from ttkbootstrap.tooltip import ToolTip
except ImportError:
    try:
        from ttkbootstrap import ToolTip
    except ImportError:
        class ToolTip:
            def __init__(self, widget, text="", bootstyle=None, **kwargs):
                pass

# Configure logging
logging.basicConfig(filename='pos.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

class ReportsManager:
    def __init__(self, app, parent, db):
        self.app = app
        self.db = db
        self.reports_frame = ttk.Frame(parent, padding=10)
        self.create_reports_tab()

    def create_reports_tab(self):
        """Create the reports interface with modern styling"""
        # Header section
        header_frame = ttk.Frame(self.reports_frame)
        header_frame.pack(fill='x', pady=(0, 20))
        
        title = ttk.Label(header_frame, text="Analytics & Reports", 
                         font=("Helvetica", 24, "bold"), foreground="#007bff")
        title.pack(side='left')
        
        # Current date display
        current_date = datetime.now().strftime("%B %d, %Y")
        date_label = ttk.Label(header_frame, text=f"Generated on {current_date}", 
                              font=("Helvetica", 12), foreground="#6c757d")
        date_label.pack(side='right')

        # Main container with two columns
        main_container = ttk.Frame(self.reports_frame)
        main_container.pack(fill='both', expand=True)
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=0)  # scrollbar column
        main_container.columnconfigure(2, weight=2)  # right column

        # Left column - Report controls with scrollbar
        left_canvas = tk.Canvas(main_container, highlightthickness=0)
        left_scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=left_canvas.yview)
        left_scrollable_frame = ttk.Frame(left_canvas)
        
        left_scrollable_frame.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        )
        
        left_canvas.create_window((0, 0), window=left_scrollable_frame, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        
        left_canvas.grid(row=0, column=0, padx=(0, 10), sticky="nsew")
        left_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Use scrollable frame as left column
        left_column = left_scrollable_frame
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            left_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        left_canvas.bind("<MouseWheel>", _on_mousewheel)

        # Quick Reports card
        quick_reports_card = create_card_frame(left_column, " Quick Reports")
        quick_reports_card.pack(fill='x', pady=(0, 15))

        # Quick report buttons with modern styling
        quick_buttons = [
            ("Current Stock", "success", self.show_stock_report),
            ("Inventory Changes", "info", self.show_inventory_adjustments),
            ("Today's Sales", "primary", self.show_today_sales),
            ("This Month", "warning", self.show_current_month_sales),
            ("This Year", "danger", self.show_current_year_sales)
        ]

        for text, style, command in quick_buttons:
            btn = ttk.Button(quick_reports_card, text=text, bootstyle=style, command=command)
            btn.pack(fill='x', pady=5, padx=10)

        # Custom Reports card
        custom_card = create_card_frame(left_column, "Custom Reports")
        custom_card.pack(fill='x', pady=(0, 15))

        # Daily Sales Report section
        daily_section = ttk.Frame(custom_card)
        daily_section.pack(fill='x', pady=10, padx=10)
        
        ttk.Label(daily_section, text="Daily Sales Report", 
                 font=("Helvetica", 12, "bold"), foreground="#495057").pack(anchor='w', pady=(0, 5))
        
        daily_input_frame = ttk.Frame(daily_section)
        daily_input_frame.pack(fill='x', pady=(0, 10))
        
        self.daily_date_var = tk.StringVar()
        self.daily_date_entry = ttk.Entry(daily_input_frame, textvariable=self.daily_date_var, 
                                         font=("Helvetica", 11))
        self.daily_date_entry.pack(fill='x', pady=(0, 5))
        self.daily_date_entry.insert(0, "YYYY-MM-DD")
        self.daily_date_entry.bind("<FocusIn>", lambda e: self.clear_placeholder(self.daily_date_entry, "YYYY-MM-DD"))
        self.daily_date_entry.bind("<FocusOut>", lambda e: self.set_placeholder(self.daily_date_entry, "YYYY-MM-DD"))
        
        daily_btn_frame = ttk.Frame(daily_input_frame)
        daily_btn_frame.pack(fill='x')
        
        generate_daily_btn = ttk.Button(daily_btn_frame, text=" Generate", 
                                       bootstyle="success", command=self.show_daily_sales)
        generate_daily_btn.pack(side='left', padx=(0, 5), fill='x', expand=True)
        
        export_daily_btn = ttk.Button(daily_btn_frame, text=" Export", 
                                     bootstyle="outline-primary", command=lambda: self.choose_export_format(self.export_daily_sales))
        export_daily_btn.pack(side='right', padx=(5, 0))

        # Monthly Sales Report section
        monthly_section = ttk.Frame(custom_card)
        monthly_section.pack(fill='x', pady=10, padx=10)
        
        ttk.Label(monthly_section, text="Monthly Sales Report", 
                 font=("Helvetica", 12, "bold"), foreground="#495057").pack(anchor='w', pady=(0, 5))
        
        monthly_input_frame = ttk.Frame(monthly_section)
        monthly_input_frame.pack(fill='x', pady=(0, 10))
        
        self.monthly_date_var = tk.StringVar()
        self.monthly_date_entry = ttk.Entry(monthly_input_frame, textvariable=self.monthly_date_var, 
                                           font=("Helvetica", 11))
        self.monthly_date_entry.pack(fill='x', pady=(0, 5))
        self.monthly_date_entry.insert(0, "YYYY-MM")
        self.monthly_date_entry.bind("<FocusIn>", lambda e: self.clear_placeholder(self.monthly_date_entry, "YYYY-MM"))
        self.monthly_date_entry.bind("<FocusOut>", lambda e: self.set_placeholder(self.monthly_date_entry, "YYYY-MM"))
        
        monthly_btn_frame = ttk.Frame(monthly_input_frame)
        monthly_btn_frame.pack(fill='x')
        
        generate_monthly_btn = ttk.Button(monthly_btn_frame, text=" Generate", 
                                         bootstyle="success", command=self.show_monthly_sales)
        generate_monthly_btn.pack(side='left', padx=(0, 5), fill='x', expand=True)
        
        export_monthly_btn = ttk.Button(monthly_btn_frame, text=" Export", 
                                       bootstyle="outline-primary", command=lambda: self.choose_export_format(self.export_monthly_sales))
        export_monthly_btn.pack(side='right', padx=(5, 0))

        # Yearly Sales Report section
        yearly_section = ttk.Frame(custom_card)
        yearly_section.pack(fill='x', pady=10, padx=10)
        
        ttk.Label(yearly_section, text="Yearly Sales Report", 
                 font=("Helvetica", 12, "bold"), foreground="#495057").pack(anchor='w', pady=(0, 5))
        
        yearly_input_frame = ttk.Frame(yearly_section)
        yearly_input_frame.pack(fill='x', pady=(0, 10))
        
        self.yearly_date_var = tk.StringVar()
        self.yearly_date_entry = ttk.Entry(yearly_input_frame, textvariable=self.yearly_date_var, 
                                          font=("Helvetica", 11))
        self.yearly_date_entry.pack(fill='x', pady=(0, 5))
        self.yearly_date_entry.insert(0, "YYYY")
        self.yearly_date_entry.bind("<FocusIn>", lambda e: self.clear_placeholder(self.yearly_date_entry, "YYYY"))
        self.yearly_date_entry.bind("<FocusOut>", lambda e: self.set_placeholder(self.yearly_date_entry, "YYYY"))
        
        yearly_btn_frame = ttk.Frame(yearly_input_frame)
        yearly_btn_frame.pack(fill='x')
        
        generate_yearly_btn = ttk.Button(yearly_btn_frame, text=" Generate", 
                                        bootstyle="success", command=self.show_yearly_sales)
        generate_yearly_btn.pack(side='left', padx=(0, 5), fill='x', expand=True)
        
        export_yearly_btn = ttk.Button(yearly_btn_frame, text=" Export", 
                                      bootstyle="outline-primary", command=lambda: self.choose_export_format(self.export_yearly_sales))
        export_yearly_btn.pack(side='right', padx=(5, 0))

        # Export All section
        export_section = ttk.Frame(custom_card)
        export_section.pack(fill='x', pady=10, padx=10)
        
        ttk.Label(export_section, text="Export Options", 
                 font=("Helvetica", 12, "bold"), foreground="#495057").pack(anchor='w', pady=(0, 5))
        
        export_stock_btn = ttk.Button(export_section, text=" Export Stock Report", 
                                     bootstyle="outline-info", command=lambda: self.choose_export_format(self.export_stock_report))
        export_stock_btn.pack(fill='x', pady=2)
        
        export_adj_btn = ttk.Button(export_section, text=" Export Adjustments", 
                                   bootstyle="outline-warning", command=lambda: self.choose_export_format(self.export_inventory_adjustments))
        export_adj_btn.pack(fill='x', pady=2)

        # Right column - Report display
        right_column = ttk.Frame(main_container)
        right_column.grid(row=0, column=2, padx=(15, 0), sticky="nsew")

        # Report display card
        self.report_card = create_card_frame(right_column, " Report Results")
        self.report_card.pack(fill='both', expand=True)

        # Report summary frame
        self.summary_frame = ttk.Frame(self.report_card)
        self.summary_frame.pack(fill='x', padx=10, pady=(10, 5))
        
        self.summary_label = ttk.Label(self.summary_frame, text="Select a report to view results", 
                                      font=("Helvetica", 12), foreground="#6c757d")
        self.summary_label.pack(anchor='w')

        # Enhanced treeview with modern styling
        tree_frame = ttk.Frame(self.report_card)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

        self.report_tree = ttk.Treeview(tree_frame, show='headings', height=20, selectmode='extended')
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.report_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.report_tree.xview)
        self.report_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid layout for tree and scrollbars
        self.report_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Configure main container row weight
        main_container.rowconfigure(0, weight=1)

        # Initialize with stock report
        self.show_stock_report()
        logging.info("Modern reports tab UI initialized")

    def clear_placeholder(self, entry, placeholder, show=None):
        """Clear placeholder text when entry is focused"""
        if entry.get() == placeholder:
            entry.delete(0, tk.END)
            if show:
                entry.configure(show=show)
        logging.info(f"Cleared placeholder for {entry}")

    def set_placeholder(self, entry, placeholder, show=None):
        """Set placeholder text when entry loses focus"""
        if not entry.get():
            entry.insert(0, placeholder)
            if show:
                entry.configure(show="")
        logging.info(f"Set placeholder for {entry}")

    def animate_cards(self, cards):
        """Apply fade-in animation to cards"""
        logging.info("Card animation skipped (not supported)")

    def show_daily_sales(self):
        """Display daily sales report"""
        try:
            date_str = self.daily_date_var.get()
            if date_str == "YYYY-MM-DD":
                messagebox.showerror("Error", "Please enter a valid date")
                logging.warning("Daily sales report failed: Placeholder date detected")
                return
            datetime.strptime(date_str, '%Y-%m-%d')  # Validate date format
            self.report_tree['columns'] = ('Product', 'Quantity Sold', 'Total (GH₵)')
            for col in self.report_tree['columns']:
                self.report_tree.heading(col, text=col)
                self.report_tree.column(col, width=150)
            
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            for row in self.db.get_daily_sales(date_str):
                product_name, total_qty, total_amount = row
                self.report_tree.insert("", "end", values=(
                    product_name, total_qty, f"GH₵{total_amount:.2f}"
                ))
            self.report_card.title_label.configure(text=f"Daily Sales Report ({date_str})")
            
            # Calculate summary statistics
            total_sales = len([row for row in self.db.get_daily_sales(date_str)])
            total_revenue = sum(row[2] for row in self.db.get_daily_sales(date_str))
            total_quantity = sum(row[1] for row in self.db.get_daily_sales(date_str))
            
            self.summary_label.configure(text=f"Sales: {total_sales} | Revenue: GH₵{total_revenue:.2f} | Items Sold: {total_quantity}")
            logging.info(f"Daily sales report generated for {date_str}")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
            logging.warning(f"Daily sales report failed: Invalid date {date_str}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate daily sales report: {str(e)}")
            logging.error(f"Daily sales report error: {str(e)}")

    def show_monthly_sales(self):
        """Display monthly sales report"""
        try:
            month_str = self.monthly_date_var.get()
            if month_str == "YYYY-MM":
                messagebox.showerror("Error", "Please enter a valid month")
                logging.warning("Monthly sales report failed: Placeholder month detected")
                return
            datetime.strptime(month_str, '%Y-%m')  # Validate month format
            self.report_tree['columns'] = ('Product', 'Category', 'Quantity Sold', 'Total (GH₵)')
            for col in self.report_tree['columns']:
                self.report_tree.heading(col, text=col)
                self.report_tree.column(col, width=150)
            
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            for row in self.db.get_monthly_sales(month_str):
                product_name, category, total_qty, total_amount = row
                self.report_tree.insert("", "end", values=(
                    product_name, category, total_qty, f"GH₵{total_amount:.2f}"
                ))
            self.report_card.title_label.configure(text=f"Monthly Sales Report ({month_str})")
            
            # Calculate summary statistics
            total_sales = len([row for row in self.db.get_monthly_sales(month_str)])
            total_revenue = sum(row[3] for row in self.db.get_monthly_sales(month_str))
            total_quantity = sum(row[2] for row in self.db.get_monthly_sales(month_str))
            
            self.summary_label.configure(text=f"Sales: {total_sales} | Revenue: GH₵{total_revenue:.2f} | Items Sold: {total_quantity}")
            logging.info(f"Monthly sales report generated for {month_str}")
        except ValueError:
            messagebox.showerror("Error", "Invalid month format. Use YYYY-MM")
            logging.warning(f"Monthly sales report failed: Invalid month {month_str}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate monthly sales report: {str(e)}")
            logging.error(f"Monthly sales report error: {str(e)}")

    def show_stock_report(self):
        """Display stock report"""
        try:
            self.report_tree['columns'] = ('Product', 'Category', 'Type', 'Unit Price (GH₵)', 'Stock', 'Stock Value (GH₵)')
            for col in self.report_tree['columns']:
                self.report_tree.heading(col, text=col)
                self.report_tree.column(col, width=120)
            
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            for row in self.db.get_stock_report():
                name, category, ptype, unit_price, stock, stock_value = row
                item_id = self.report_tree.insert("", "end", values=(
                    name, category, ptype, f"GH₵{unit_price:.2f}", stock, f"GH₵{stock_value:.2f}"
                ))
                if (category == 'Block' and stock < 10) or (category == 'Cement' and stock < 5):
                    self.report_tree.item(item_id, tags=('low_stock',))
            self.report_tree.tag_configure('low_stock', foreground='red')
            self.report_card.title_label.configure(text="Current Stock Report")
            
            # Calculate summary statistics
            total_products = len([row for row in self.db.get_stock_report()])
            total_value = sum(row[5] for row in self.db.get_stock_report())
            low_stock_items = len([row for row in self.db.get_stock_report() 
                                 if (row[1] == 'Block' and row[4] < 10) or (row[1] == 'Cement' and row[4] < 5)])
            
            self.summary_label.configure(text=f"Products: {total_products} | Total Value: GH₵{total_value:.2f} | Low Stock: {low_stock_items}")
            logging.info("Stock report generated")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate stock report: {str(e)}")
            logging.error(f"Stock report error: {str(e)}")

    def show_inventory_adjustments(self):
        """Display inventory adjustments report"""
        try:
            self.report_tree['columns'] = ('ID', 'Product', 'Quantity Change', 'Note', 'Date')
            for col in self.report_tree['columns']:
                self.report_tree.heading(col, text=col)
                self.report_tree.column(col, width=150)
            
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            total_adjustments = 0
            for row in self.db.get_inventory_logs():
                log_id, product_name, change_qty, note, log_date = row
                formatted_date = datetime.fromisoformat(log_date).strftime("%Y-%m-%d %H:%M:%S")
                self.report_tree.insert("", "end", values=(
                    log_id, product_name, change_qty, note or "", formatted_date
                ))
                total_adjustments += 1
            
            self.report_card.title_label.configure(text="Inventory Adjustments Report")
            self.summary_label.configure(text=f"Total adjustments: {total_adjustments}")
            logging.info("Inventory adjustments report generated")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate inventory adjustments report: {str(e)}")
            logging.error(f"Inventory adjustments report error: {str(e)}")

    def show_today_sales(self):
        """Display today's sales report"""
        today = datetime.now().strftime('%Y-%m-%d')
        self.daily_date_var.set(today)
        self.show_daily_sales()

    def show_current_month_sales(self):
        """Display current month's sales report"""
        current_month = datetime.now().strftime('%Y-%m')
        self.monthly_date_var.set(current_month)
        self.show_monthly_sales()

    def show_current_year_sales(self):
        """Display current year's sales report"""
        current_year = datetime.now().strftime('%Y')
        self.yearly_date_var.set(current_year)
        self.show_yearly_sales()

    def show_yearly_sales(self):
        """Display yearly sales report"""
        try:
            year_str = self.yearly_date_var.get()
            if year_str == "YYYY":
                messagebox.showerror("Error", "Please enter a valid year")
                logging.warning("Yearly sales report failed: Placeholder year detected")
                return
            
            # Validate year format
            try:
                year = int(year_str)
                if year < 1900 or year > 2100:
                    raise ValueError("Year out of range")
            except ValueError:
                messagebox.showerror("Error", "Invalid year format. Use YYYY")
                logging.warning(f"Yearly sales report failed: Invalid year {year_str}")
                return
                
            self.report_tree['columns'] = ('Product', 'Category', 'Quantity Sold', 'Total (GH₵)')
            for col in self.report_tree['columns']:
                self.report_tree.heading(col, text=col)
                self.report_tree.column(col, width=150)
            
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            for row in self.db.get_yearly_sales(year_str):
                product_name, category, total_qty, total_amount = row
                self.report_tree.insert("", "end", values=(
                    product_name, category, total_qty, f"GH₵{total_amount:.2f}"
                ))
            self.report_card.title_label.configure(text=f"Yearly Sales Report ({year_str})")
            
            # Calculate summary statistics
            total_sales = len([row for row in self.db.get_yearly_sales(year_str)])
            total_revenue = sum(row[3] for row in self.db.get_yearly_sales(year_str))
            total_quantity = sum(row[2] for row in self.db.get_yearly_sales(year_str))
            
            self.summary_label.configure(text=f"Sales: {total_sales} | Revenue: GH₵{total_revenue:.2f} | Items Sold: {total_quantity}")
            logging.info(f"Yearly sales report generated for {year_str}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate yearly sales report: {str(e)}")
            logging.error(f"Yearly sales report error: {str(e)}")

    def export_daily_sales(self, format_type='csv'):
        """Export daily sales report to CSV or Excel"""
        try:
            date_str = self.daily_date_var.get()
            if date_str == "YYYY-MM-DD":
                messagebox.showerror("Error", "Please enter a valid date")
                logging.warning("Daily sales export failed: Placeholder date detected")
                return
            datetime.strptime(date_str, '%Y-%m-%d')  # Validate date format
            
            headers = ['Product', 'Quantity Sold', 'Total (GH₵)']
            data = []
            for row in self.db.get_daily_sales(date_str):
                product_name, total_qty, total_amount = row
                if format_type == 'excel':
                    data.append([product_name, total_qty, f"GH₵{total_amount:.2f}"])
                else:
                    data.append([product_name, total_qty, f"{total_amount:.2f}"])
            
            if format_type == 'excel' and EXCEL_AVAILABLE:
                filename = f"daily_sales_{date_str}.xlsx"
                if self.create_excel_workbook(f"Daily Sales Report - {date_str}", headers, data, filename):
                    messagebox.showinfo("Success", f"Daily sales report exported to {filename}")
                    logging.info(f"Daily sales report exported to {filename}")
                else:
                    messagebox.showerror("Error", "Failed to create Excel file")
            else:
                filename = f"daily_sales_{date_str}.csv"
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for row in data:
                        if format_type == 'csv':
                            writer.writerow([row[0], row[1], f"{float(row[2].replace('GH₵', '')) if isinstance(row[2], str) and row[2].startswith('GH₵') else row[2]:.2f}"])
                        else:
                            writer.writerow(row)
                messagebox.showinfo("Success", f"Daily sales report exported to {filename}")
                logging.info(f"Daily sales report exported to {filename}")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
            logging.warning(f"Daily sales export failed: Invalid date {date_str}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export daily sales: {str(e)}")
            logging.error(f"Daily sales export error: {str(e)}")

    def export_monthly_sales(self, format_type='csv'):
        """Export monthly sales report to CSV or Excel"""
        try:
            month_str = self.monthly_date_var.get()
            if month_str == "YYYY-MM":
                messagebox.showerror("Error", "Please enter a valid month")
                logging.warning("Monthly sales export failed: Placeholder month detected")
                return
            datetime.strptime(month_str, '%Y-%m')  # Validate month format
            
            headers = ['Product', 'Category', 'Quantity Sold', 'Total (GH₵)']
            data = []
            for row in self.db.get_monthly_sales(month_str):
                product_name, category, total_qty, total_amount = row
                if format_type == 'excel':
                    data.append([product_name, category, total_qty, f"GH₵{total_amount:.2f}"])
                else:
                    data.append([product_name, category, total_qty, f"{total_amount:.2f}"])
            
            if format_type == 'excel' and EXCEL_AVAILABLE:
                filename = f"monthly_sales_{month_str}.xlsx"
                if self.create_excel_workbook(f"Monthly Sales Report - {month_str}", headers, data, filename):
                    messagebox.showinfo("Success", f"Monthly sales report exported to {filename}")
                    logging.info(f"Monthly sales report exported to {filename}")
                else:
                    messagebox.showerror("Error", "Failed to create Excel file")
            else:
                filename = f"monthly_sales_{month_str}.csv"
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for row in data:
                        if format_type == 'csv':
                            writer.writerow([row[0], row[1], row[2], f"{float(row[3].replace('GH₵', '')) if isinstance(row[3], str) and row[3].startswith('GH₵') else row[3]:.2f}"])
                        else:
                            writer.writerow(row)
                messagebox.showinfo("Success", f"Monthly sales report exported to {filename}")
                logging.info(f"Monthly sales report exported to {filename}")
        except ValueError:
            messagebox.showerror("Error", "Invalid month format. Use YYYY-MM")
            logging.warning(f"Monthly sales export failed: Invalid month {month_str}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export monthly sales: {str(e)}")
            logging.error(f"Monthly sales export error: {str(e)}")

    def export_stock_report(self, format_type='csv'):
        """Export stock report to CSV or Excel"""
        try:
            headers = ['Product', 'Category', 'Type', 'Unit Price (GH₵)', 'Stock', 'Stock Value (GH₵)']
            data = []
            for row in self.db.get_stock_report():
                name, category, ptype, unit_price, stock, stock_value = row
                if format_type == 'excel':
                    data.append([name, category, ptype, f"GH₵{unit_price:.2f}", stock, f"GH₵{stock_value:.2f}"])
                else:
                    data.append([name, category, ptype, f"{unit_price:.2f}", stock, f"{stock_value:.2f}"])
            
            if format_type == 'excel' and EXCEL_AVAILABLE:
                filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                if self.create_excel_workbook("Current Stock Report", headers, data, filename):
                    messagebox.showinfo("Success", f"Stock report exported to {filename}")
                    logging.info(f"Stock report exported to {filename}")
                else:
                    messagebox.showerror("Error", "Failed to create Excel file")
            else:
                filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(data)
                messagebox.showinfo("Success", f"Stock report exported to {filename}")
                logging.info(f"Stock report exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export stock report: {str(e)}")
            logging.error(f"Stock report export error: {str(e)}")

    def export_inventory_adjustments(self, format_type='csv'):
        """Export inventory adjustments to CSV or Excel"""
        try:
            headers = ['ID', 'Product', 'Quantity Change', 'Note', 'Date']
            data = []
            for row in self.db.get_inventory_logs():
                log_id, product_name, change_qty, note, log_date = row
                formatted_date = datetime.fromisoformat(log_date).strftime("%Y-%m-%d %H:%M:%S")
                data.append([log_id, product_name, change_qty, note or "", formatted_date])
            
            if format_type == 'excel' and EXCEL_AVAILABLE:
                filename = f"inventory_adjustments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                if self.create_excel_workbook("Inventory Adjustments Report", headers, data, filename):
                    messagebox.showinfo("Success", f"Inventory adjustments exported to {filename}")
                    logging.info(f"Inventory adjustments exported to {filename}")
                else:
                    messagebox.showerror("Error", "Failed to create Excel file")
            else:
                filename = f"inventory_adjustments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(data)
                messagebox.showinfo("Success", f"Inventory adjustments exported to {filename}")
                logging.info(f"Inventory adjustments exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export inventory adjustments: {str(e)}")
            logging.error(f"Inventory adjustments export error: {str(e)}")

    def export_yearly_sales(self, format_type='csv'):
        """Export yearly sales report to CSV or Excel"""
        try:
            year_str = self.yearly_date_var.get()
            if year_str == "YYYY":
                messagebox.showerror("Error", "Please enter a valid year")
                logging.warning("Yearly sales export failed: Placeholder year detected")
                return
            
            # Validate year format
            try:
                year = int(year_str)
                if year < 1900 or year > 2100:
                    raise ValueError("Year out of range")
            except ValueError:
                messagebox.showerror("Error", "Invalid year format. Use YYYY")
                logging.warning(f"Yearly sales export failed: Invalid year {year_str}")
                return
            
            headers = ['Product', 'Category', 'Quantity Sold', 'Total (GH₵)']
            data = []
            for row in self.db.get_yearly_sales(year_str):
                product_name, category, total_qty, total_amount = row
                if format_type == 'excel':
                    data.append([product_name, category, total_qty, f"GH₵{total_amount:.2f}"])
                else:
                    data.append([product_name, category, total_qty, f"{total_amount:.2f}"])
            
            if format_type == 'excel' and EXCEL_AVAILABLE:
                filename = f"yearly_sales_{year_str}.xlsx"
                if self.create_excel_workbook(f"Yearly Sales Report - {year_str}", headers, data, filename):
                    messagebox.showinfo("Success", f"Yearly sales report exported to {filename}")
                    logging.info(f"Yearly sales report exported to {filename}")
                else:
                    messagebox.showerror("Error", "Failed to create Excel file")
            else:
                filename = f"yearly_sales_{year_str}.csv"
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for row in data:
                        if format_type == 'csv':
                            writer.writerow([row[0], row[1], row[2], f"{float(row[3].replace('GH₵', '')) if isinstance(row[3], str) and row[3].startswith('GH₵') else row[3]:.2f}"])
                        else:
                            writer.writerow(row)
                messagebox.showinfo("Success", f"Yearly sales report exported to {filename}")
                logging.info(f"Yearly sales report exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export yearly sales: {str(e)}")
            logging.error(f"Yearly sales export error: {str(e)}")

    def refresh_reports(self):
        """Refresh the current report display"""
        try:
            current_label = self.report_card.title_label.cget('text')
            if "Daily Sales" in current_label:
                self.show_daily_sales()
            elif "Monthly Sales" in current_label:
                self.show_monthly_sales()
            elif "Stock Report" in current_label:
                self.show_stock_report()
            elif "Inventory Adjustments" in current_label:
                self.show_inventory_adjustments()
            elif "Yearly Sales" in current_label:
                self.show_yearly_sales()
            logging.info("Reports refreshed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh reports: {str(e)}")
            logging.error(f"Error refreshing reports: {str(e)}")

    def choose_export_format(self, export_function, *args):
        """Allow user to choose between CSV and Excel export formats"""
        if not EXCEL_AVAILABLE:
            export_function(*args, format_type='csv')
            return
            
        # Create a dialog to choose format
        format_dialog = tk.Toplevel()
        format_dialog.title("Choose Export Format")
        format_dialog.geometry("300x150")
        format_dialog.transient(self.reports_frame.winfo_toplevel())
        format_dialog.grab_set()
        
        # Center the dialog
        format_dialog.update_idletasks()
        x = (format_dialog.winfo_screenwidth() // 2) - (300 // 2)
        y = (format_dialog.winfo_screenheight() // 2) - (150 // 2)
        format_dialog.geometry(f"300x150+{x}+{y}")
        
        ttk.Label(format_dialog, text="Select export format:", font=("Helvetica", 12)).pack(pady=20)
        
        button_frame = ttk.Frame(format_dialog)
        button_frame.pack(pady=10)
        
        def export_csv():
            format_dialog.destroy()
            export_function(*args, format_type='csv')
            
        def export_excel():
            format_dialog.destroy()
            export_function(*args, format_type='excel')
        
        ttk.Button(button_frame, text="CSV", bootstyle="outline-primary", 
                  command=export_csv).pack(side='left', padx=10)
        ttk.Button(button_frame, text="Excel", bootstyle="primary", 
                  command=export_excel).pack(side='right', padx=10)

    def create_excel_workbook(self, title, headers, data, filename):
        """Create a professionally formatted Excel workbook"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
            
            # Set up styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add title
            ws.merge_cells('A1:' + get_column_letter(len(headers)) + '1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal="center")
            
            # Add generation date
            ws.merge_cells('A2:' + get_column_letter(len(headers)) + '2')
            date_cell = ws['A2']
            date_cell.value = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_cell.alignment = Alignment(horizontal="center")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            for row_idx, row_data in enumerate(data, 5):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Clean currency formatting for Excel
                    if isinstance(value, str) and value.startswith('GH₵'):
                        cell.value = float(value.replace('GH₵', '').replace(',', ''))
                        cell.number_format = '"GH₵"#,##0.00'
                    else:
                        cell.value = value
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            return True
        except Exception as e:
            logging.error(f"Error creating Excel workbook: {str(e)}")
            return False